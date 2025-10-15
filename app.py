import streamlit as st
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl.styles import Font
import time
import os

# === Streamlit UI ===
st.title("📱 Flipkart Smartphone Scraper")
st.markdown("This app scrapes top 20 smartphone names and prices from Flipkart and exports them to an Excel file.")

if st.button("Scrape Flipkart"):
    with st.spinner("Scraping data from Flipkart..."):

        # === Selenium Setup ===
        # You must have msedgedriver in your system PATH or provide full path below
        options = webdriver.EdgeOptions()
        options.add_argument('--headless')  # Run in headless mode (no GUI)
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920x1080')

        try:
            driver = webdriver.Edge(options=options)

            # Navigate to Flipkart
            driver.get("https://www.flipkart.com/")
            time.sleep(2)

            # Close login popup if present
            try:
                close_btn = driver.find_element(By.XPATH, "//button[contains(text(),'✕')]")
                close_btn.click()
            except:
                pass  # If popup doesn't appear, skip

            # Search for smartphones
            search_box = driver.find_element(By.NAME, "q")
            search_box.send_keys("Smartphones")
            search_box.send_keys(Keys.RETURN)

            time.sleep(5)  # Wait for results to load

            # Get product names and prices
            product_names = driver.find_elements(By.CSS_SELECTOR, "div.KzDlHZ")[:20]
            product_prices = driver.find_elements(By.CSS_SELECTOR, "div.Nx9bqj._4b5DiR")[:20]

            smartphone_name = [p.text for p in product_names]
            smartphone_price = [p.text for p in product_prices]

            driver.quit()

            # === Excel File Creation ===
            wb = Workbook()
            ws = wb.active
            ws.title = "Top20 Smartphone Results"

            ws['A1'] = "Smartphone Name"
            ws['B1'] = "Smartphone Price"

            # Styling
            header_font = Font(bold=True, size=14)
            ws["A1"].font = header_font
            ws["B1"].font = header_font

            # Writing data
            for i in range(len(smartphone_name)):
                ws.cell(row=i+2, column=1).value = smartphone_name[i]
                ws.cell(row=i+2, column=2).value = smartphone_price[i]

            # Save to file
            filename = "smartphone.xlsx"
            wb.save(filename)

            st.success("✅ Scraping completed and data saved to Excel!")
            with open(filename, "rb") as f:
                st.download_button(
                    label="📥 Download Excel F
